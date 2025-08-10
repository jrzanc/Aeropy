from openpyxl import load_workbook
from datetime import datetime, timedelta
import re

# Classe Spreadsheet
class SpreadsheetMetadata:
    def __init__(self,address):
        self.address = address
        self.cabecalho=self.get_cabecalho(self.address)
        self.cia=self.set_cia()
        self.start_billing_period=self.set_start_billing_period()
        self.end_billing_period=self.set_end_billing_period()
        
    def get_cabecalho(self, num_linhas=6):
        ws = load_workbook(filename=self.address, data_only=True).active
        cabecalho=[row for row in ws.iter_rows(min_row=1, max_row=6, values_only=True)]
        return cabecalho

    def set_cia(self):
        return self.cabecalho[1][2]
        
    def set_start_billing_period(self):
        date=re.split(r"[ /]",self.cabecalho[1][5])
        start_str="-".join([date[0],date[1],date[-1]])+" 00:00:00"        
        start = datetime.strptime(start_str, "%d-%m-%Y %H:%M:%S")        
        return start
        
    def set_end_billing_period(self):
        date=re.split(r"[ /]",self.cabecalho[1][5])
        end_str="-".join([date[-3],date[-2],date[-1]])+" 23:59:59"      
        end=datetime.strptime(end_str, "%d-%m-%Y %H:%M:%S")+timedelta(seconds=1)      
        return end   
