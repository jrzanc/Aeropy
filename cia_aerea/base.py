
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re

# Importar a classe SpreadsheetMetadata, assumindo que ela está em outro arquivo

from planilhas.metadata import SpreadsheetMetadata


class CiaAerea:
    def __init__(self,address):
        self.address=address
        self.metadata=SpreadsheetMetadata(address)
        self.table=self.get_table()  
        
    def get_sheet_name(self):
        return None
        
    def get_table(self):
        name = self.get_sheet_name()
        tb=pd.read_excel(self.address,sheet_name=name,skiprows=6)
        return tb

    def _loc_teste(self,name_df,columns,ty):
        df = getattr(self, name_df)
        for col in columns:
            df[col] = df[col].astype(ty)
        setattr(self,name_df,df)


    def data_structure(self):
        keys_cols={"string":["Localizador",
                             "Bilhete", "Órgão Solicitante",
                             "Órgão Superior",
                             "PCDP" ,
                             "Evento(s) de Faturamento",
                             "Situação do Bilhete",
                             "Situações de Embarque",
                             "Status do Faturamento",
                             "Tipo(s) de Pendência",
                             "Análise da Fiscalização",
                             "Análise da Companhia Aérea"],
                  "datetime64[ns]":["Data de emissão",
                                    "Data de Cancelamento"],
                  "float64":['Crédito de Reembolso',
                             'Desconto %',
                             'Desconto Efetivo %',
                             'Desconto Efetivo R$',
                             'Desconto R$',
                             'Glosa',
                             'Multa Calculada',
                             'Multa Efetiva',
                             'Multa de Reembolso Retornada',
                             'Reembolso Calculado',
                             'Reembolso Efetivo',
                             'Reembolso Retornado',
                             'Tarifa Comercial Retornada',
                             'Tarifa Embarque Efetiva',
                             'Tarifa Praticada Calculada',
                             'Tarifa Praticada Efetiva',
                             'Tarifa Praticada Retornada',
                             'Tarifa de Embarque Retornada',
                             'Total',
                             'Valor a Faturar']
                  }
        for key in keys_cols.keys():
            self._loc_teste("table",keys_cols[key],key)

    def insert_message_on_tickets_with_discount_bellow_three_percent(self):
        df = getattr(self, "table") 
        filter_discount_below_three_percent= df[df["Tipo(s) de Pendência"].str.contains("Pendente de Desconto", na=False)]
        new_text="Validar o valor a faturar - Não foi aplicado o desconto - bilhete glosado"
        df.loc[filter_discount_below_three_percent.index,"Análise da Fiscalização"]=(df.loc[filter_discount_below_three_percent.index,"Análise da Fiscalização"]
                                                                                        .apply(lambda x: new_text if pd.isna(x) else f"{x}\n{new_text}"
                                                                                                                                                                        )
                                                                                        )
        setattr(self,"table",df)

    def insert_message_about_72_hours(self):
        df = getattr(self, "table") 
        filter_message_about_72_hours= df[df["Tipo(s) de Pendência"].str.contains("Pendende de Reserva", na=False)]
        
        new_text="Validar o valor a faturar - Não foi respeitado o prazo de 72 horas da reserva"
        df.loc[filter_message_about_72_hours.index,"Análise da Fiscalização"]=(df.loc[filter_message_about_72_hours.index,"Análise da Fiscalização"]
                                                                                        .apply(lambda x: new_text if pd.isna(x) else f"{x}\n{new_text}"
                                                                                                                                                                        )
                                                                                        )
        setattr(self,"table",df)

      
    def general_message(self):
        df = getattr(self, "table") 
        df.loc[df["Análise da Fiscalização"].isna(),"Análise da Fiscalização"]="Validar o valor a faturar"
        setattr(self,"table",df)


    def insert_message_on_tickets_canceled_after_the_billing_period(self):
        df = getattr(self, "table") 
        #tbl=self.tickets_issued.copy()
        filter_billing_period=df[(~df['Data de Cancelamento'].isnull()
                                )&(df['Data de Cancelamento']>=self.metadata.end_billing_period)
                                ]
        new_text="Cancelado no mês seguinte ao de referência. Será reembolsado no próximo faturamento."
        
        df.loc[filter_billing_period.index, "Análise da Fiscalização"] = (df.loc[filter_billing_period.index, "Análise da Fiscalização"]
                                                                            .apply(lambda x: new_text if pd.isna(x) else f"{x}\n{new_text}"
                                                                                    )
                                                                            )
        setattr(self,"table",df)


    def change_worksheet(self):
        wb = load_workbook(self.address)
        ws = wb[self.get_sheet_name()]

        # Encontrar a coluna onde está "Análise da Fiscalização" na linha 8
        header_row = 7
        target_col = None
        for cell in ws[header_row]:
            if cell.value == "Análise da Fiscalização":
                target_col = cell.column  # Isso retorna o número da coluna (não a letra)
                break

        if target_col is None:
            raise ValueError("Coluna 'Análise da Fiscalização' não encontrada na linha 8.")

        # Escrever os valores na coluna correta
        for i, valor in enumerate(self.table["Análise da Fiscalização"], start=8):
            ws.cell(row=i, column=target_col, value=valor)

        wb.save(self.address)

'''
    def change_worksheet(self):
        wb = load_workbook(self.address)
        ws = wb[self.get_sheet_name()] 
        for i, valor in enumerate(self.table["Análise da Fiscalização"], start=8):
            ws.cell(row=i, column=33, value=valor)
        wb.save(self.address)   '''