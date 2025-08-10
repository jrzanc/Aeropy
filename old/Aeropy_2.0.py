
# Importing Labraries
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re
import os
import numpy as np
import sys
from PySide6.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QVBoxLayout,
    QHBoxLayout, QGroupBox, QMessageBox, QRadioButton, QButtonGroup
)
from PySide6.QtGui import QPixmap
from PySide6.QtCore import Qt
import sys


# classe planilhas
class planilha:
    def get_address(self):
        # Cria uma janela oculta
        root = tk.Tk()
        root.withdraw()
        # Abre a janela para selecionar o arquivo
        try:
            caminho_arquivo = filedialog.askopenfilename(
                title="Selecione um arquivo",
                filetypes=[("Arquivos Excel", "*.xls"),("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
            )
            return caminho_arquivo
        except Exception as error:
            print(f"error type:{error}")
            
    def get_new_address(self,address): 
        address_splited=address.split("/")
        old_name_splited=address_splited[-1].split(".")
        new_name=old_name_splited[0]+"_MODIFIED_SPREADSHEET.xlsx"
        new_address = "/".join(address_splited[:-1]) + "/" + new_name
        return new_address

    def copy_spreadsheet(self,old_address,new_address):
        with open(old_address, 'rb') as origem:
            conteudo = origem.read()
        try:
            with open(new_address, 'wb') as destino:
                destino.write(conteudo)
            print("Arquivo Excel copiado com sucesso!")
        except Exception as error:
            print(f"error type:{error}")
   
    def flow(self):
        g_address=self.get_address()
        n_address=self.get_new_address(g_address)
        self.copy_spreadsheet(g_address,n_address)
        return n_address
# Classe Spreadsheet
class SpreadsheetMetadata:
    def __init__(self,address):
        self.address = address
        self.cabecalho=self.get_cabecalho(self.address)
        self.cia=self.set_cia()
        self.start_billing_period=self.set_start_billing_period()
        self.end_billing_period=self.set_end_billing_period()
        
    def get_cabecalho(self, num_linhas=6):
        ws = load_workbook(filename=self.address, data_only=True).active
        cabecalho=[row for row in ws.iter_rows(min_row=1, max_row=6, values_only=True)]
        return cabecalho

    def set_cia(self):
        return self.cabecalho[1][2]
        
    def set_start_billing_period(self):
        date=re.split(r"[ /]",self.cabecalho[1][5])
        start_str="-".join([date[0],date[1],date[-1]])+" 00:00:00"        
        start = datetime.strptime(start_str, "%d-%m-%Y %H:%M:%S")        
        return start
        
    def set_end_billing_period(self):
        date=re.split(r"[ /]",self.cabecalho[1][5])
        end_str="-".join([date[-3],date[-2],date[-1]])+" 23:59:59"      
        end=datetime.strptime(end_str, "%d-%m-%Y %H:%M:%S")+timedelta(seconds=1)      
        return end   

# Classe Cia Aérea
class CiaAerea:
    def __init__(self,address):
        self.address=address
        self.metadata=SpreadsheetMetadata(address)
        self.table=self.get_table()  
        
    def get_sheet_name(self):
        return None
        
    def get_table(self):
        name = self.get_sheet_name()
        tb=pd.read_excel(self.address,sheet_name=name,skiprows=6)
        return tb

    def _loc_teste(self,name_df,columns,ty):
        df = getattr(self, name_df)
        for col in columns:
            df[col] = df[col].astype(ty)
        setattr(self,name_df,df)


    def data_structure(self):
        keys_cols={"string":["Localizador",
                             "Bilhete", "Órgão Solicitante",
                             "Órgão Superior",
                             "PCDP" ,
                             "Evento(s) de Faturamento",
                             "Situação do Bilhete",
                             "Situações de Embarque",
                             "Status do Faturamento",
                             "Tipo(s) de Pendência",
                             "Análise da Fiscalização",
                             "Análise da Companhia Aérea"],
                  "datetime64[ns]":["Data de emissão",
                                    "Data de Cancelamento"],
                  "float64":['Crédito de Reembolso',
                             'Desconto %',
                             'Desconto Efetivo %',
                             'Desconto Efetivo R$',
                             'Desconto R$',
                             'Glosa',
                             'Multa Calculada',
                             'Multa Efetiva',
                             'Multa de Reembolso Retornada',
                             'Reembolso Calculado',
                             'Reembolso Efetivo',
                             'Reembolso Retornado',
                             'Tarifa Comercial Retornada',
                             'Tarifa Embarque Efetiva',
                             'Tarifa Praticada Calculada',
                             'Tarifa Praticada Efetiva',
                             'Tarifa Praticada Retornada',
                             'Tarifa de Embarque Retornada',
                             'Total',
                             'Valor a Faturar']
                  }
        for key in keys_cols.keys():
            self._loc_teste("table",keys_cols[key],key)

    def insert_message_on_tickets_with_discount_bellow_three_percent(self):
        df = getattr(self, "table") 
        filter_discount_below_three_percent=df[df['Desconto %']<0.03]
        new_text="Validar o valor a faturar - Não foi aplicado o desconto - bilhete glosado"
        df.loc[filter_discount_below_three_percent.index,"Análise da Fiscalização"]=(df.loc[filter_discount_below_three_percent.index,"Análise da Fiscalização"]
                                                                                      .apply(lambda x: new_text if pd.isna(x) else f"{x}\n{new_text}")
                                                                                    )
                          

    def general_message(self):
        df = getattr(self, "table") 
        df.loc[df["Análise da Fiscalização"].isna(),"Análise da Fiscalização"]="Validar o valor a faturar"
        setattr(self,"table",df)


    def change_worksheet(self):
        wb = load_workbook(self.address)
        ws = wb[self.get_sheet_name()] 
        for i, valor in enumerate(self.table["Análise da Fiscalização"], start=8):
            ws.cell(row=i, column=33, value=valor)
        wb.save(self.address)   

# Classe Cia Aérea Emitidos
class CiaAreaIssued(CiaAerea):
    def __init__(self,address):
        super().__init__(address)
        self.table=self.get_table()   

    def get_sheet_name(self):
        return "Emitidos"

    def insert_message_on_tickets_canceled_after_the_billing_period(self):
        df = getattr(self, "table") 
        #tbl=self.tickets_issued.copy()
        filter_billing_period=df[(~df['Data de Cancelamento'].isnull()
                                )&(df['Data de Cancelamento']>=self.metadata.start_billing_period
                                  )&(df['Data de Cancelamento']<self.metadata.end_billing_period)
                                ]
        new_text="Cancelado no mês seguinte ao de referência. Será reembolsado no próximo faturamento."
        
        df.loc[filter_billing_period.index, "Análise da Fiscalização"] = (df.loc[filter_billing_period.index, "Análise da Fiscalização"]
                                                                           .apply(lambda x: new_text if pd.isna(x) else f"{x}\n{new_text}"
                                                                                 )
                                                                          )
        setattr(self,"table",df)

    def insert_message_about_72_hours(self):
        df = getattr(self, "table") 
        filter_message_about_72_hours= df[df["Tipo(s) de Pendência"].str.contains("Pendende de Reserva", na=False)]
        
        new_text="Validar o valor a faturar - Não foi respeitado o prazo de 72 horas da reserva"
        df.loc[filter_message_about_72_hours.index,"Análise da Fiscalização"]=(df.loc[filter_message_about_72_hours.index,"Análise da Fiscalização"]
                                                                                      .apply(lambda x: new_text if pd.isna(x) else f"{x}\n{new_text}"
                                                                                                                                                                        )
                                                                                     )
        setattr(self,"table",df)
        
    def flow(self):
        self.data_structure()
        self.insert_message_on_tickets_canceled_after_the_billing_period()
        self.insert_message_on_tickets_with_discount_bellow_three_percent()
        self.insert_message_about_72_hours()
        self.general_message()
        self.change_worksheet()


# Classe emitidos e Cancelados  
class CiaAreaIssuedCancelled(CiaAerea):
    def __init__(self,address):
        super().__init__(address)
        self.table=self.get_table()   

    def get_sheet_name(self):
        return "Emitidos e Cancelados"
        
    def void_tickets(self):
        df = getattr(self, "table") 
        df.loc[df["Status do Faturamento"] == "Não Faturável", "Análise da Fiscalização"] = "VOID"
        setattr(self,"table",df)

    def insert_message_on_tickets_with_discount_bellow_three_percent(self):
        df = getattr(self, "table") 
        filter_discount_below_three_percent=df[df['Desconto %']<0.03]
        new_text="Validar o valor do reembolso indicado na coluna AA - Não foi aplicado o desconto - bilhete glosado"
        df.loc[filter_discount_below_three_percent.index,"Análise da Fiscalização"]=(df.loc[filter_discount_below_three_percent.index,"Análise da Fiscalização"]
                                                                                      .apply(lambda x: new_text if pd.isna(x) else f"{x}\n{new_text}"
                                                                                            )
                                                                                    )

    def general_message(self):
        df = getattr(self, "table") 
        df.loc[df["Análise da Fiscalização"].isna(),"Análise da Fiscalização"]="Validar o valor do reembolso indicado na coluna AA"
        setattr(self,"table",df)    
    
    
    def flow(self):
        self.data_structure()
        self.void_tickets()
        self.insert_message_on_tickets_with_discount_bellow_three_percent()
        self.general_message()
        self.change_worksheet()

# Classe Visual
class TelaInicial(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Aeropy - Central de Compras")
        self.setGeometry(100, 100, 300, 200)
        self.setStyleSheet("background-color: #007b8a; color: white; font-family: Arial;")

        # Logo JPG à esquerda
        logo = QLabel()
        pixmap = QPixmap("logo.jpg")
        pixmap = pixmap.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        logo.setPixmap(pixmap)
        logo.setFixedSize(100, 100)

        # Título e subtítulo à direita da logo
        titulo = QLabel("Bem-vindo ao Aeropy")
        titulo.setStyleSheet("font-size: 20px; font-weight: bold;")
        subtitulo = QLabel("Criado por Junior A. Zancanaro")
        subtitulo.setStyleSheet("font-size: 12px;")

        titulo_layout = QVBoxLayout()
        titulo_layout.addWidget(titulo, alignment=Qt.AlignCenter)
        titulo_layout.addWidget(subtitulo, alignment=Qt.AlignCenter)

        # Layout horizontal para logo + título
        topo_layout = QHBoxLayout()
        topo_layout.addWidget(logo)
        topo_layout.addLayout(titulo_layout)

        # Instrução
        instrucao = QLabel("Selecione qual parte do processo de medição você está executando:")
        instrucao.setAlignment(Qt.AlignCenter)

        # Radio buttons
        self.radio_group = QButtonGroup(self)
        self.radio_provisoria = QRadioButton("Etapa de medição provisória")
        self.radio_definitiva = QRadioButton("Etapa de medição definitiva")
        for rb in [self.radio_provisoria, self.radio_definitiva]:
            rb.setStyleSheet("font-size: 16px;")
            self.radio_group.addButton(rb)

        # Agrupando os radio buttons
        box = QGroupBox("Etapas disponíveis")
        box.setStyleSheet("QGroupBox { font-size: 18px; font-weight: bold; }")
        box_layout = QVBoxLayout()
        box_layout.addWidget(self.radio_provisoria)
        box_layout.addWidget(self.radio_definitiva)
        box.setLayout(box_layout)

        # Botões
        btn_continuar = QPushButton("Continuar")
        btn_cancelar = QPushButton("Cancelar")
        btn_continuar.clicked.connect(self.continuar)
        btn_cancelar.clicked.connect(self.fechar_aplicacao)

        for btn in [btn_continuar, btn_cancelar]:
            btn.setStyleSheet("background-color: white; color: #007b8a; font-weight: bold; padding: 8px;")

        # Layout principal
        layout = QVBoxLayout()
        layout.addLayout(topo_layout)
        layout.addWidget(instrucao)
        layout.addWidget(box)

        botoes_layout = QHBoxLayout()
        botoes_layout.addWidget(btn_continuar)
        botoes_layout.addWidget(btn_cancelar)
        layout.addLayout(botoes_layout)

        self.setLayout(layout)

    def continuar(self):
        if self.radio_provisoria.isChecked():
            try:
                worksheet_excel = planilha()
                address = worksheet_excel.flow()
                issued = CiaAreaIssued(address)
                issued.flow()
                issued_cancelled = CiaAreaIssuedCancelled(address)
                issued_cancelled.flow()
                QMessageBox.information(self, "Sucesso", "Processo executado com sucesso!")
            except:
                QMessageBox.critical(self, "Erro", "Erro ao manusear a planilha!")
            finally:
                self.close()
                QApplication.instance().quit()

        elif self.radio_definitiva.isChecked():
            QMessageBox.information(self, "Continuar", "Iniciando a etapa: Etapa de medição definitiva")

        else:
            QMessageBox.warning(self, "Aviso", "Por favor, selecione uma etapa antes de continuar.")

    def fechar_aplicacao(self):
        self.close()
        QApplication.instance().quit()


# Método Main
if __name__ == "__main__":
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    janela = TelaInicial()
    janela.show()
    sys.exit(app.exec())